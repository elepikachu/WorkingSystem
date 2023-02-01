from django.http import HttpResponse, HttpResponseRedirect
from django.utils.encoding import escape_uri_path
from openpyxl.utils import get_column_letter
from .models import Item, ItemLog, Suggestion
from django.core.paginator import Paginator
from django.shortcuts import render
from openpyxl.styles import Font
from bs4 import BeautifulSoup
from io import BytesIO

import pandas as pd
import numpy as np

import datetime
import requests
import json


VERSION = '物资采购系统 1.0.0'


# -------------------------------------------------------------
# 函数名： get_ip
# 功能： 获取电脑主机ip
# -------------------------------------------------------------
def get_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# -------------------------------------------------------------
# 函数名： project_view
# 功能： 项目管理系统首页
# -------------------------------------------------------------
def buy_view(request):
    dic = {'ver': VERSION}
    return render(request, 'buyitem/buy.html', dic)


# -------------------------------------------------------------
# 函数名： suggest_view
# 功能： 意见建议页面
# -------------------------------------------------------------
def suggest_view(request):
    if request.method == 'GET':
        dic = {'ver': VERSION}
        return render(request, 'buyitem/buysuggestion.html', dic)
    elif request.method == 'POST':
        if 'sub' in request.POST:
            sug = request.POST['sug']
            typ = request.POST['typ']
            if Suggestion.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            Suggestion.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), detail=sug, type=typ, finish=0)
            return HttpResponseRedirect('/buyitem')


# -------------------------------------------------------------
# 函数名： submit_view
# 功能： 提交模式
# -------------------------------------------------------------
def submit_view(request):
    if request.method == 'GET':
        if request.COOKIES.get('name', '') != '':
            psn = request.COOKIES.get('name', '')
            psn = json.loads(psn)
            grp = request.COOKIES.get('group', '')
            grp = json.loads(grp)
            tel = request.COOKIES.get('tel', '')
            tel = json.loads(tel)
            num = request.COOKIES.get('num', '')
            num = json.loads(num)
        else:
            psn = ''
            grp = ''
            tel = ''
            num = ''
        if 'good' in request.GET:
            god = request.GET['good']
            no = request.GET['no']
            shp = request.GET['shop']
            if no != '':
                no = '商品编号：' + no
        else:
            god = ''
            no = ''
            shp = ''
        dic = {'ver': VERSION, 'psn': psn, 'grp': grp, 'tel': tel, 'num': num, 'god': god, 'no': no, 'shp': shp}
        return render(request, 'buyitem/buysubmit.html', dic)
    elif request.method == 'POST':
        if 'sub' in request.POST:
            name = request.POST['name']
            group = request.POST['group']
            tel = request.POST['phone']
            num = request.POST['num']
            good = request.POST['good']
            brand = request.POST['brand']
            quantity = request.POST['quantity']
            unit = request.POST['unit']
            info = request.POST['info']
            cif = request.POST['cif']
            detail = request.POST['detail']
            if Item.objects.exists():
                id = Item.objects.latest('id').id + 1
            else:
                id = 1
            Item.objects.create(id=id, name=name, phone=tel, group=group, num=num, unit=unit, date=datetime.datetime.today(),
                                detail=detail, quantity=quantity, brand=brand, info=info, finish=0, good=good, classif=cif)
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='insert',
                                   other='%s-%s-%s-%s' % (name, good, num, 'false'))
            res = HttpResponseRedirect('/buyitem/manage?page=1')
            name = json.dumps(name)
            group = json.dumps(group)
            tel = json.dumps(tel)
            num = json.dumps(num)

            res.set_cookie(key='name', value=name, max_age=3600 * 24 * 30)
            res.set_cookie(key='group', value=group, max_age=3600 * 24 * 30)
            res.set_cookie(key='tel', value=tel, max_age=3600 * 24 * 30)
            res.set_cookie(key='num', value=num, max_age=3600 * 24 * 30)
            return res
        elif 'spd' in request.POST:
            if request.POST['good'] == '':
                return HttpResponse('商品为空')
            rep = HttpResponseRedirect('/buyitem/spider?good=%s&pg=1' % request.POST['good'])
            if request.COOKIES.get('name') == '':
                name = request.POST['name']
                group = request.POST['group']
                tel = request.POST['phone']
                num = request.POST['num']
                rep.set_cookie(key='name', value=name, max_age=3600 * 24 * 30)
                rep.set_cookie(key='group', value=group, max_age=3600 * 24 * 30)
                rep.set_cookie(key='tel', value=tel, max_age=3600 * 24 * 30)
                rep.set_cookie(key='num', value=num, max_age=3600 * 24 * 30)
            return rep


# -------------------------------------------------------------
# 函数名： spider_view
# 功能： 爬取数据展示
# -------------------------------------------------------------
def spider_view(request):
    page = request.GET['pg']
    good = request.GET['good']
    res = parse_page(good, page)
    page = int(page)
    dic = {'ver': VERSION, 'res': res, 'good': good, 'lastpg': page-1, 'nextpg': page+1}
    return render(request, 'buyitem/buyspider.html', dic)


# -------------------------------------------------------------
# 函数名： manage_view
# 功能： 管理模式
# -------------------------------------------------------------
def manage_view(request):
    if request.method == 'GET':
        if 'grp' in request.GET:
            if request.GET['grp'] is "全部" and request.GET['fin'] is 2 or '2':
                all_data = Item.objects.filter(date__gte=request.GET['st'], date__lte=request.GET['et'])
            elif request.GET['fin'] is 2 or '2':
                all_data = Item.objects.filter(date__gte=request.GET['st'], date__lte=request.GET['et'], group__exact=request.GET['grp'])
            elif request.GET['grp'] is "全部":
                all_data = Item.objects.filter(date__gte=request.GET['st'], date__lte=request.GET['et'], finish__exact=request.GET['fin'])
            else:
                all_data = Item.objects.filter(date__gte=request.GET['st'], date__lte=request.GET['et'], finish__exact=request.GET['fin'], group__exact=request.GET['grp'])
            if request.GET['fin'] == 0:
                slo = "未完成"
            elif request.GET['fin'] == 1:
                slo = "已完成"
            else:
                slo = ""
        else:
            all_data = Item.objects.filter(finish__exact=False)
            all_data = all_data[::-1]
            slo = "未完成"
        page_num = request.GET.get('page', 1)
        paginator = Paginator(all_data, 30)
        c_page = paginator.page(int(page_num))
        ver = VERSION
        date = datetime.date.today()
        datee = date - datetime.timedelta(weeks=4)
        date1 = datee.strftime('%Y-%m-%d')
        date2 = date.strftime('%Y-%m-%d')
        groupbox = []
        for item in all_data:
            if item.group not in groupbox:
                groupbox.append(item.group)
        groupbox.append('全部')
        return render(request, 'buyitem/buymanage.html', locals())
    elif request.method == 'POST':
        if 'del' in request.POST:
            return HttpResponseRedirect('/buyitem/batch')
        if 'self' in request.POST:
            return HttpResponseRedirect('/buyitem/personal')
        if 'pri' in request.POST:
            mon = (datetime.datetime.today()).strftime("%m")
            mon = str(int(mon) + 1)
            if request.POST['grp'] == "全部" and request.POST['fin'] == "全部":
                all_data = Item.objects.filter(date__gte=request.POST['date1'], date__lte=request.POST['date2'])
                name = f'物资采购信息-全.xlsx'
                ofc = False
            elif request.POST['fin'] == "全部":
                all_data = Item.objects.filter(date__gte=request.POST['date1'], date__lte=request.POST['date2'],
                                                  group__exact=request.POST['grp'])
                name = f'物资采购信息-全-%s.xlsx' % request.POST['grp']
                ofc = True
            elif request.POST['grp'] == "全部":
                if request.POST['fin'] == "未完成":
                    fin = 0
                    name = f'股份月度计划-%s月.xlsx' % mon
                elif request.POST['fin'] == "已完成":
                    fin = 1
                    name = f'物资采购信息-已完成.xlsx'
                all_data = Item.objects.filter(date__gte=request.POST['date1'], date__lte=request.POST['date2'],
                                               finish__exact=fin)
                ofc = False
            else:
                if request.POST['fin'] == "未完成":
                    fin = 0
                    name = f'股份月度计划-%s月-%s.xlsx' % (mon, request.POST['grp'])
                elif request.POST['fin'] == "已完成":
                    fin = 1
                    name = f'物资采购信息-已完成-%s.xlsx' % request.POST['grp']
                all_data = Item.objects.filter(date__gte=request.POST['date1'], date__lte=request.POST['date2'],
                                               finish__exact=fin, group__exact=request.POST['grp'])
                ofc = True
            data_list = all_data.values_list()
            response = create_excel(data_list, name, ofc)
            return response
        if 'unfit' in request.POST:
            return HttpResponseRedirect('/buyitem/manage?page=1')
        if 'fit' in request.POST:
            st = request.POST['date1']
            et = request.POST['date2']
            grp = request.POST['grp']
            if request.POST['fin'] == "全部":
                fin = 2
            elif request.POST['fin'] == "未完成":
                fin = 0
            elif request.POST['fin'] == "已完成":
                fin = 1
            return HttpResponseRedirect('/buyitem/manage?page=1&st=%s&et=%s&grp=%s&fin=%d' % (st, et, grp, fin))


# -------------------------------------------------------------
# 函数名： update_buy
# 功能： 数据行更新
# -------------------------------------------------------------
def update_buy(request, item_id):
    try:
        item = Item.objects.get(id=item_id)

    except Exception as e:
        print('--update error is %s' % e)
        return HttpResponse('--The project is not existed!!--')
    if request.method == 'GET':
        ver = VERSION
        return render(request, 'buyitem/buyupdate.html', locals())
    elif request.method == 'POST':
        if "upd" in request.POST:
            name = request.POST['name']
            group = request.POST['group']
            tel = request.POST['phone']
            num = request.POST['num']
            good = request.POST['good']
            brand = request.POST['brand']
            quantity = request.POST['quantity']
            unit = request.POST['unit']
            info = request.POST['info']
            detail = request.POST['detail']
            cif = request.POST['cif']
            #finish = request.POST['finish']
            item.name = name
            item.tel = tel
            item.group = group
            item.num = num
            item.good = good
            item.detail = detail
            item.brand = brand
            item.quantity = quantity
            item.unit = unit
            item.classif = cif
            item.info = info
            #item.finish = finish
            item.save()
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='update',
                                   other='%s-%s-%s-%s' % (name, good, num, item.finish))
            return HttpResponseRedirect('/buyitem/manage?page=1')
        elif "back" in request.POST:
            return HttpResponseRedirect('/buyitem/manage?page=1')


# -------------------------------------------------------------
# 函数名： finish_buy
# 功能： 数据行完成
# -------------------------------------------------------------
def finish_buy(request, item_id):
    try:
        item = Item.objects.get(id=item_id)
        if item.finish == True:
            return HttpResponse('--操作失败，该条购买已经完成啦!!--')
        item.finish = True
        item.save()
        name = item.name
        good = item.good
        if ItemLog.objects.exists():
            idx = ItemLog.objects.latest('id').id
        else:
            idx = 0
        ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='finish',
                               other='%s-%s' % (name, good))
        return HttpResponseRedirect('/buyitem/manage?page=1')
    except Exception as e:
        print('--delete error is %s' % e)
        return HttpResponse('--更新失败，请稍后再试!!--')


# -------------------------------------------------------------
# 函数名： delete_buy
# 功能： 数据行删除
# -------------------------------------------------------------
def delete_buy(request, item_id):
    try:
        item = Item.objects.get(id=item_id)
        name = item.name
        good = item.good
        item.delete()
        if ItemLog.objects.exists():
            idx = ItemLog.objects.latest('id').id
        else:
            idx = 0
        ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='delete',
                               other='%s-%s' % (name, good))
        return HttpResponseRedirect('/buyitem/manage?page=1')
    except Exception as e:
        print('--delete error is %s' % e)
        return HttpResponse('--删除失败，请稍后再试!!--')


# -------------------------------------------------------------
# 函数名： batch_view
# 功能： 批量删除
# -------------------------------------------------------------
def batch_view(request):
    if request.method == 'GET':
        date = datetime.date.today()
        datee = date + datetime.timedelta(weeks=1)
        date1 = date.strftime('%Y-%m-%d')
        date2 = datee.strftime('%Y-%m-%d')
        groupbox = []
        all_data = Item.objects.all()
        for item in all_data:
            if item.group not in groupbox:
                groupbox.append(item.group)
        ver = VERSION
        return render(request, 'buyitem/buydelete.html', locals())
    elif request.method == 'POST':
        if 'delall' in request.POST:
            all_data = Item.objects.all()
            all_data.delete()
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='delete',
                                   other='all')
            return HttpResponseRedirect('/buyitem/manage?page=1')
        elif 'deldate' in request.POST:
            date1 = request.POST['date1']
            date2 = request.POST['date2']
            all_data = Item.objects.filter(date__gte=date1, date__lte=date2)
            all_data.delete()
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='delete',
                                      other='all in %s-%s' % (date1, date2))
            return HttpResponseRedirect('/buyitem/manage?page=1')
        elif 'delgrp' in request.POST:
            grp = request.POST['grp']
            all_data = Item.objects.filter(group__exact=grp)
            all_data.delete()
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='delete',
                                other='all for %s' % grp)
            return HttpResponseRedirect('/buyitem/manage?page=1')
        elif 'delcif' in request.POST:
            cif = request.POST['cif']
            all_data = Item.objects.filter(classif__exact=cif)
            all_data.delete()
            if ItemLog.objects.exists():
                idx = ItemLog.objects.latest('id').id
            else:
                idx = 0
            ItemLog.objects.create(id=idx + 1, ip=get_ip(request), date=datetime.datetime.today(), cmd='delete',
                                other='all for %s' % cif)
            return HttpResponseRedirect('/buyitem/manage?page=1')


# -------------------------------------------------------------
# 函数名： personal_view
# 功能： 个人数据操作
# -------------------------------------------------------------
def personal_view(request):
    if request.method == 'GET':
        ver = VERSION
        if request.COOKIES.get('name', '') != '':
            psn = request.COOKIES.get('name', '')
            psn = json.loads(psn)
            unfinished = Item.objects.filter(name__exact=psn).exclude(finish__exact=True)
            unfinished = unfinished[::-1]
            length = len(unfinished)
            finished = Item.objects.filter(name__exact=psn, finish__exact=True)
            finished = finished[::-1]
            return render(request, 'buyitem/buyself.html', locals())
        else:
            namebox = []
            all_data = Item.objects.all()
            for item in all_data:
                if item.name not in namebox:
                    namebox.append(item.name)
            return render(request, 'buyitem/buyperson.html', locals())
    if request.method == 'POST':
        if 'ret' in request.POST:
            res = HttpResponseRedirect('/buyitem/personal')
            res.delete_cookie('name')
            return res
        if 'nm' in request.POST:
            res = HttpResponseRedirect('/buyitem/personal')
            person = json.dumps(request.POST['nam'])
            res.set_cookie(key='name', value=person, max_age=3600 * 24 * 30)
            return res
        if 'prt1' in request.POST:
            psn = request.COOKIES.get('name', '')
            psn = json.loads(psn)
            all_data = Item.objects.filter(name__exact=psn, finish__exact=0)
            data_list = all_data.values_list()
            mon = (datetime.datetime.today()).strftime("%m")
            mon = str(int(mon) + 1)
            name = f'股份月度计划-%s月-%s.xlsx' % (mon, psn)
            response = create_excel(data_list, name, True)
            return response
        if 'prt2' in request.POST:
            psn = request.COOKIES.get('name', '')
            psn = json.loads(psn)
            all_data = Item.objects.filter(name__exact=psn)
            data_list = all_data.values_list()
            mon = (datetime.datetime.today()).strftime("%m")
            mon = str(int(mon) + 1)
            name = f'股份月度计划-%s月-%s.xlsx' % (mon, psn)
            response = create_excel(data_list, name, True)
            return response


# -------------------------------------------------------------
# 函数名： create_new_excel
# 功能： 打印excel
# -------------------------------------------------------------
def create_new_excel(data_list, name):
    data = pd.DataFrame(data_list)
    data.columns = ['id', '商品名', '品牌型号', '单位', '数量', '姓名', '电话', '课题编号', '采购说明', '备注', '单位全称', '提交日期', '完成情况']
    data['备注'] = "商品编号: " + data["备注"]
    output = BytesIO()  # 转二进制流
    data.to_excel(output, index=False)
    output.seek(0)  # 重新定位到开始
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment;filename=%s" % escape_uri_path(name)
    response.write(output.getvalue())
    output.close()
    return response


# -------------------------------------------------------------
# 函数名： excel_style
# 功能： 添加excel的格式
# -------------------------------------------------------------
def excel_style(x):
    return ['text-align:center' for x in x]


# -------------------------------------------------------------
# 函数名： create_excel
# 功能： 生成excel
# -------------------------------------------------------------
def create_excel(data_list, name, office):

    rawdata = pd.DataFrame(data_list)
    rawdata.columns = ['序号', '商品名称', '品牌型号', '单位', '数量', '姓名', '电话', '课题编号', '采购说明', '备注', '11', '12', '13', '14']
    #rawdata['备注'] = "商品编号: " + rawdata["备注"]

    output = BytesIO()  # 转二进制流
    writer = pd.ExcelWriter(output, engine='openpyxl')
    writer = print_excel(writer, '股份-办公用品采购', '办公用品', rawdata, office)
    writer = print_excel(writer, '股份-设备耗材采购', '设备耗材', rawdata, office)
    writer = print_excel(writer, '股份-办公家具', '办公家具', rawdata, office)
    writer = print_excel(writer, '股份-五金杂品采购', '五金杂品', rawdata, office)
    writer = print_excel(writer, '股份-劳动防护用品', '劳动防护', rawdata, office)
    writer = print_excel(writer, '股份-实验耗材及小型设备', '实验耗材及小型设备', rawdata, office)

    writer.save()
    output.seek(0)  # 重新定位到开始
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = "attachment;filename=%s" % escape_uri_path(name)
    response.write(output.getvalue())
    output.close()
    return response


# -------------------------------------------------------------
# 函数名： print_excel
# 功能： 按类别把数据输出到excel
# -------------------------------------------------------------
def print_excel(writer, word, wd, rawdata, of):
    data = rawdata[rawdata['12'] == wd]
    if not data.empty:
        data = data.drop(['11', '12', '13', '14'], axis=1)
        data.style.apply(excel_style, axis=0).to_excel(writer, sheet_name=word, index=False, startrow=3)

        column_wid = (data.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values)
        max_wid = (data.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values)
        wids = np.max([column_wid, max_wid], axis=0)
        worksheet = writer.sheets[word]
        for i, wid in enumerate(wids, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = wid + 2
    else:
        try:
            data = data.drop(['11', '12', '13', '14'], axis=1)
        except:
            pass
        data.to_excel(writer, sheet_name=word, index=False, startrow=3)

    worksheet = writer.sheets[word]
    mon = (datetime.datetime.today()).strftime("%m")
    mon = str(int(mon) + 1)
    worksheet.cell(row=1, column=1).value = '            中国石油勘探开发研究院采购计划表(' + mon + '月)'
    worksheet['A1'].font = Font(size=20, bold=True)
    worksheet.cell(row=2, column=1).value = '表单号：NKZ0400-01'
    worksheet['A2'].font = Font(size=12, bold=True)
    worksheet.cell(row=2, column=6).value = '类别：' + wd
    worksheet['F2'].font = Font(size=12, bold=True)
    if of:
        worksheet.cell(row=3, column=1).value = '申请单位：' + rawdata['11'][0]
    else:
        worksheet.cell(row=3, column=1).value = '申请单位：'
    worksheet['A3'].font = Font(size=12, bold=True)
    worksheet.cell(row=3, column=6).value = '填报日期：' + datetime.datetime.today().strftime("%Y-%m-%d")
    worksheet['F3'].font = Font(size=12, bold=True)
    return writer


# -------------------------------------------------------------
# 函数名： log_view
# 功能： 日志界面
# -------------------------------------------------------------
def log_view(request):
    if request.method == 'GET':
        all_log = ItemLog.objects.all()
        all_log = all_log[::-1]
        dic = {'ver': VERSION, 'data': all_log}
        return render(request, 'buyitem/buylog.html', dic)


# -------------------------------------------------------------
# 函数名： get_page
# 功能： 爬取页面
# -------------------------------------------------------------
def get_page(url):
    headers = {
         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
    }
    try:
        html = requests.request("GET", url, headers=headers, timeout=10)
        html.encoding = "utf-8"
        return html.text
    except:
        print('爬取失败')
        return "爬取失败"


# -------------------------------------------------------------
# 函数名： parse_page
# 功能： 解析爬取页面
# -------------------------------------------------------------
def parse_page(item, page):
    url = "https://search.jd.com/Search?keyword=%s&page=%s" % (item, page)
    html = get_page(url)
    html = str(html)
    if html is not None:
        soup = BeautifulSoup(html, 'html.parser')
        li_all = soup.select('#J_goodsList ul li')
        res_list = []
        for li in li_all:
            name = [i.get_text() for i in li.select('.p-name em')][0]
            price = [i.get_text() for i in li.select('.p-price i')][0]
            if li.select('.p-shop a'):
                shop = [i.get_text() for i in li.select('.p-shop a')][0]
            elif li.select('.p-shopnum a'):
                shop = [i.get_text() for i in li.select('.p-shopnum a')][0]
            else:
                shop = "自营"
            number = li['data-sku']
            href = "item.jd.com/%s.html" % number
            if(len(name) !=0 and len(price) !=0 and len(shop) !=0 and len(number) !=0):
                info = {'name': name, 'price': price, 'shop': shop, 'number': number, 'href': href}
                res_list.append(info)
        return res_list
    else:
        print('error')
        return None


